from flask import Flask, render_template, request, redirect, send_file, url_for, session,flash,Response,render_template_string,send_file
import os
import openpyxl
from werkzeug.utils import secure_filename
import sqlite3
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from flask import send_from_directory
from datetime import date
from flask import jsonify
from leave_management import *
from reporting_functions import *
from performance_management import *
from datetime import datetime
from werkzeug.security import check_password_hash, generate_password_hash
import csv
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
pdfmetrics.registerFont(TTFont('Arabic', r'C:\Windows\Fonts\arial.ttf'))
import io
from weasyprint import HTML
from openpyxl import Workbook


app = Flask(__name__)
app.secret_key = "secret123"

# إعدادات رفع ملفات الـ CV
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CV_UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads", "cvs")
os.makedirs(CV_UPLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = CV_UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB كحد أقصى للملف

ALLOWED_CV_EXTENSIONS = {"pdf", "doc", "docx"}


def allowed_cv_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_CV_EXTENSIONS
# ====== إنشاء قاعدة البيانات والجداول تلقائيًا ======
def init_db():
    conn = sqlite3.connect('hr.db')
    c = conn.cursor()

    # جدول الموظفين
    c.execute('''
    CREATE TABLE IF NOT EXISTS employees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    department TEXT,
    position TEXT,
    salary REAL,
    username TEXT UNIQUE,
    password_hash TEXT,
    phone TEXT,
    email TEXT,
    document TEXT,
    basic_salary REAL
)
''')

    # جدول الحضور
    c.execute('''
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER,
        date TEXT,
        check_in TEXT,
        check_out TEXT,
        FOREIGN KEY(employee_id) REFERENCES employees(id)
    )
    ''')

    conn.commit()
    conn.close()

def ensure_employee_portal_columns():
    """إضافة أعمدة اسم المستخدم وكلمة المرور إذا كانت القاعدة قديمة"""
    conn = sqlite3.connect('hr.db')
    c = conn.cursor()
    info = c.execute("PRAGMA table_info(employees)").fetchall()
    col_names = [row[1] for row in info]
    if 'username' not in col_names:
        c.execute("ALTER TABLE employees ADD COLUMN username TEXT")
    if 'password_hash' not in col_names:
        c.execute("ALTER TABLE employees ADD COLUMN password_hash TEXT")
    conn.commit()
    conn.close()

def ensure_payroll_schema():
    """تجهيز جداول الرواتب والتوسعات المتقدمة إن لم تكن موجودة."""
    conn = sqlite3.connect('hr.db')
    c = conn.cursor()

    # جدول الرواتب 
    c.execute("""
        CREATE TABLE IF NOT EXISTS payrolls (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            salary REAL NOT NULL,
            bonus REAL DEFAULT 0,
            deductions REAL DEFAULT 0,
            net_salary REAL NOT NULL,
            payment_date TEXT NOT NULL,
            month TEXT,
            FOREIGN KEY(employee_id) REFERENCES employees(id)
        )
    """)

    # إضافة أعمدة ناقصة 
    cols = [r[1] for r in c.execute("PRAGMA table_info(payrolls)").fetchall()]
    def add_col(name, ddl):
        if name not in cols:
            c.execute(f"ALTER TABLE payrolls ADD COLUMN {ddl}")

    add_col("salary", "salary REAL")
    add_col("bonus", "bonus REAL DEFAULT 0")
    add_col("deductions", "deductions REAL DEFAULT 0")
    add_col("net_salary", "net_salary REAL")
    add_col("payment_date", "payment_date TEXT")
    add_col("month", "month TEXT")

    # بنود الرواتب المتقدمة (بدلات/خصومات متعددة لكل مسير)
    c.execute("""
        CREATE TABLE IF NOT EXISTS payroll_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_id INTEGER NOT NULL,
            item_type TEXT NOT NULL, -- allowance | deduction
            label TEXT NOT NULL,
            amount REAL NOT NULL,
            FOREIGN KEY(payroll_id) REFERENCES payrolls(id)
        )
    """)

    conn.commit()
    conn.close()

# تشغيل الإنشاء
init_db()
ensure_employee_portal_columns()
ensure_payroll_schema()

# باقي الكود الخاص بـ Flask هنا...

app.secret_key = 'secretkey'

def get_db_connection():
    conn = sqlite3.connect('hr.db')
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] == 'admin' and request.form['password'] == 'admin':
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            error = 'بيانات الدخول غير صحيحة'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    total_employees = conn.execute("SELECT COUNT(*) AS c FROM employees").fetchone()["c"]
    today_str = date.today().strftime("%Y-%m-%d")
    present_today = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) AS c FROM attendance WHERE date = ? AND check_in IS NOT NULL",
        (today_str,)
    ).fetchone()["c"]
    absent_today = max(total_employees - present_today, 0)
    conn.close()
    return render_template(
        'index.html',
        total_employees_home=total_employees,
        today_date_home=today_str,
        present_today_home=present_today,
        absent_today_home=absent_today,
    )

@app.route('/employees')
def employees():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    keyword = request.args.get('keyword', '')
    try:
        per_page = int(request.args.get('per_page', 20))
        if per_page not in (10, 20, 50):
            per_page = 20
    except ValueError:
        per_page = 20

    # إعداد التصفح (Pagination)
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1

    offset = (page - 1) * per_page

    conn = get_db_connection()
    if keyword:
        employees = conn.execute(
            "SELECT * FROM employees WHERE name LIKE ? LIMIT ? OFFSET ?",
            ('%' + keyword + '%', per_page, offset)
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM employees WHERE name LIKE ?",
            ('%' + keyword + '%',)
        ).fetchone()['c']
    else:
        employees = conn.execute(
            "SELECT * FROM employees LIMIT ? OFFSET ?",
            (per_page, offset)
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM employees"
        ).fetchone()['c']

    conn.close()

    pages = (total + per_page - 1) // per_page if total else 1

    return render_template(
        'employee_list.html',
        employees=employees,
        keyword=keyword,
        page=page,
        pages=pages,
        total=total,
        per_page=per_page
    )

@app.route('/employee/<int:emp_id>')
def employee_details(emp_id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    employee = conn.execute("SELECT * FROM employees WHERE id = ?", (emp_id,)).fetchone()
    attendance = conn.execute("SELECT * FROM attendance WHERE employee_id = ?", (emp_id,)).fetchall()
    conn.close()
    return render_template('employee_details.html', employee=employee, attendance=attendance)


@app.route('/employees/cvs')
def employee_cvs():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    employees = conn.execute(
        "SELECT id, name, department, position, document FROM employees WHERE document IS NOT NULL AND document != ''"
    ).fetchall()
    conn.close()
    return render_template('employee_cvs.html', employees=employees)


@app.route('/employees/cv/<int:employee_id>')
def download_employee_cv(employee_id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    row = conn.execute(
        "SELECT name, document FROM employees WHERE id = ?", (employee_id,)
    ).fetchone()
    conn.close()
    if not row or not row["document"]:
        flash("لا يوجد ملف CV محفوظ لهذا الموظف.", "error")
        return redirect(url_for('employee_cvs'))
    return send_from_directory(app.config["UPLOAD_FOLDER"], row["document"], as_attachment=True)

@app.route('/add', methods=['GET', 'POST'])
def add_employee():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            department = request.form.get('department', '').strip()
            position = request.form.get('position', '').strip()
            phone = request.form.get('phone', '').strip()
            email = request.form.get('email', '').strip()
            address=request.form.get('address', '').strip()
            if not name or not department or not position:
                flash("الاسم والقسم والوظيفة مطلوبة.", "error")
                return redirect(url_for("add_employee"))
            try:
                salary = float(request.form.get('salary', 0) or 0)
            except (TypeError, ValueError):
                flash("أدخل راتباً رقمياً صحيحاً.", "error")
                return redirect(url_for("add_employee"))
        except Exception as e:
            flash(f"خطأ في البيانات المدخلة: {str(e)}", "error")
            return redirect(url_for("add_employee"))

        portal_username = request.form.get('portal_username', '').strip()
        portal_password = request.form.get('portal_password', '')

        # معالجة ملف الـ CV (اختياري)
        cv_file = request.files.get('cv_file') or request.files.get('cv')
        cv_filename = None
        if cv_file and cv_file.filename:
            if allowed_cv_file(cv_file.filename):
                safe_name = secure_filename(cv_file.filename)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                cv_filename = f"{timestamp}_{safe_name}"
                cv_file.save(os.path.join(app.config["UPLOAD_FOLDER"], cv_filename))
            else:
                flash("صيغة ملف الـ CV غير مدعومة. المسموح: pdf, doc, docx", "error")
                return redirect(url_for("add_employee"))

        username_val = None
        password_hash_val = None
        if portal_username and portal_password:
            username_val = portal_username
            password_hash_val = generate_password_hash(portal_password)

        conn = get_db_connection()
        try:
            conn.execute(
                """INSERT INTO employees (name, department, position, salary, phone, email, document, username, password_hash, address)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, department, position, salary, phone, email, cv_filename, username_val, password_hash_val,address)
            )
            conn.commit()
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("اسم المستخدم للبوابة مُستخدم مسبقاً أو يوجد موظف بنفس الاسم. غيّر الاسم أو اسم المستخدم.", "error")
            return redirect(url_for("add_employee"))
        finally:
            conn.close()

        return redirect(url_for("employees"))
    return render_template("add_employee.html")

@app.route('/delete/<int:id>')
def delete_employee(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    conn.execute("DELETE FROM employees WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return redirect('/employees')

@app.route('/attendance')
def attendance():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    records = conn.execute('''
        SELECT a.id, e.name, a.date, a.check_in, a.check_out
        FROM attendance a
        JOIN employees e ON a.employee_id = e.id
        ORDER BY a.date DESC
    ''').fetchall()
    conn.close()
    return render_template('attendance.html', records=records)

@app.route('/attendance/checkin/<int:employee_id>', methods=['POST'])
def check_in(employee_id):
    today = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')
    conn = get_db_connection()
    record = conn.execute("SELECT * FROM attendance WHERE employee_id = ? AND date = ?", (employee_id, today)).fetchone()
    if not record:
        conn.execute("INSERT INTO attendance (employee_id, date, check_in) VALUES (?, ?, ?)",
                     (employee_id, today, time_now))
        conn.commit()
    conn.close()
    return redirect('/attendance')

@app.route('/attendance/checkout/<int:employee_id>', methods=['POST'])
def check_out(employee_id):
    today = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')
    conn = get_db_connection()
    conn.execute("UPDATE attendance SET check_out = ? WHERE employee_id = ? AND date = ?",
                 (time_now, employee_id, today))
    conn.commit()
    conn.close()
    return redirect('/attendance')

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_employee(id):
    conn = get_db_connection()

    if request.method == 'POST':
        name = request.form['name']
        department = request.form['department']
        position = request.form['position']
        salary = request.form['salary']
        portal_username = request.form.get('portal_username', '').strip()
        portal_password = request.form.get('portal_password', '')

        try:
            if portal_password:
                password_hash_val = generate_password_hash(portal_password)
                conn.execute(
                    """UPDATE employees SET name=?, department=?, position=?, salary=?, username=?, password_hash=? WHERE id=?""",
                    (name, department, position, salary, portal_username or None, password_hash_val, id)
                )
            else:
                conn.execute(
                    """UPDATE employees SET name=?, department=?, position=?, salary=?, username=? WHERE id=?""",
                    (name, department, position, salary, portal_username or None, id)
                )
            conn.commit()
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("اسم المستخدم للبوابة مُستخدم من موظف آخر. اختر اسماً مختلفاً.", "error")
            conn.close()
            return redirect(url_for('edit_employee', id=id))
        finally:
            conn.close()
        return redirect(url_for('employees'))

    employee = conn.execute("SELECT * FROM employees WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template('edit_employee.html', employee=employee)

@app.route('/download_report')
def download_report():
      if not session.get('logged_in'):
          return redirect(url_for('login'))

      conn = get_db_connection()
      data = conn.execute('SELECT * FROM employees ORDER BY name').fetchall()
      conn.close()

      # Create Excel workbook
      wb = Workbook()
      ws = wb.active
      ws.title = "Employees"

      # Headers (Arabic)
      headers = ['ID', 'الاسم' ,'الوظيفه' ,'القسم' ,'الراتب','العنوان']
      ws.append(headers)

      # Style header row
      for cell in ws[1]:
          cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
          cell.fill = openpyxl.styles.PatternFill(start_color="0b1c8c", end_color="0b1c8c", fill_type="solid")

      # Add data rows
      for row in data:
          ws.append([
    row["id"],
    row["name"],
    row["department"],
    row["position"],
    row["salary"],
    row["address"]
])

      # Save to BytesIO
      output = io.BytesIO()
      wb.save(output)
      output.seek(0)

      return send_file(
          output,
          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
          as_attachment=True,
          download_name='employee_report.xlsx'
      )
@app.route('/employee/login', methods=['GET', 'POST'])
def employee_login():
    """تسجيل دخول الموظف للبوابة الذاتية"""
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db_connection()
        employee = conn.execute(
            "SELECT * FROM employees WHERE username = ?", (username,)
        ).fetchone()
        conn.close()

        if employee and employee['password_hash']:
            if check_password_hash(employee['password_hash'], password):
                session.clear()
                session['employee_logged_in'] = True
                session['employee_id'] = employee['id']
                session['employee_name'] = employee['name']
                return redirect(url_for('self_portal'))
        if employee and not employee['password_hash']:
            error = 'حسابك غير مفعّل للبوابة الذاتية. تواصل مع المسؤول لتعيين اسم مستخدم وكلمة مرور.'
        else:
            error = 'بيانات الدخول غير صحيحة'

    return render_template('employee_login.html', error=error)


@app.route('/employee/logout')
def employee_logout():
    """تسجيل خروج الموظف من البوابة الذاتية"""
    session.pop('employee_logged_in', None)
    session.pop('employee_id', None)
    session.pop('employee_name', None)
    return redirect(url_for('employee_login'))


@app.route('/self/portal')
def self_portal():
    """الصفحة الرئيسية للبوابة الذاتية للموظف"""
    if not session.get('employee_logged_in'):
        return redirect(url_for('employee_login'))

    emp_id = session.get('employee_id')
    conn = get_db_connection()
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (emp_id,)).fetchone()
    conn.close()

    return render_template('self_portal.html', employee=employee)


@app.route('/self/leaves/request', methods=['GET', 'POST'])
def self_request_leave():
    """طلب إجازة من الموظف نفسه + عرض رصيده"""
    if not session.get('employee_logged_in'):
        return redirect(url_for('employee_login'))

    employee_id = session.get('employee_id')

    if request.method == 'POST':
        leave_type_id = request.form['leave_type_id']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        reason = request.form['reason']

        success, message = submit_leave_request(employee_id, leave_type_id, start_date, end_date, reason)

        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')

        return redirect(url_for('self_request_leave'))

    conn = get_db_connection()
    leave_types = conn.execute('SELECT id, name, max_days FROM leave_types').fetchall()
    conn.close()

    balances = get_employee_leave_balances(employee_id)

    return render_template(
        'self_request_leave.html',
        leave_types=leave_types,
        balances=balances,
    )


@app.route('/self/performance/history')
def self_performance_history():
    """تاريخ أداء الموظف (من منظور الموظف نفسه)"""
    if not session.get('employee_logged_in'):
        return redirect(url_for('employee_login'))

    employee_id = session.get('employee_id')

    conn = get_db_connection()
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    conn.close()

    if not employee:
        flash('الموظف غير موجود', 'error')
        return redirect(url_for('self_portal'))

    evaluations = get_employee_evaluations(employee_id)
    trend = get_employee_performance_trend(employee_id)

    return render_template(
        'employee_performance_history.html',
        employee=employee,
        evaluations=evaluations,
        trend=trend,
    )


@app.route('/self/payroll')
def self_payroll():
    """عرض رواتب الموظف وقسائم الرواتب"""
    if not session.get('employee_logged_in'):
        return redirect(url_for('employee_login'))

    employee_id = session.get('employee_id')

    conn = sqlite3.connect('hr.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        '''
        SELECT id, salary, bonus, deductions, net_salary, payment_date
        FROM payrolls
        WHERE employee_id = ?
        ORDER BY payment_date DESC
        ''',
        (employee_id,),
    )
    payrolls = cur.fetchall()
    conn.close()

    return render_template('self_payroll.html', payrolls=payrolls)
@app.route('/dashboard')
def dashboard():
    conn = get_db_connection()

    # عدد الموظفين
    total_employees = conn.execute('SELECT COUNT(*) as count FROM employees').fetchone()['count']

    # عدد الأقسام
    total_departments = conn.execute('SELECT COUNT(DISTINCT department) as count FROM employees').fetchone()['count']

    # عدد الحضور اليوم (مثال إذا عندك جدول attendance)
    total_present = conn.execute(
        "SELECT COUNT(DISTINCT employee_id) as count FROM attendance WHERE date = date('now')"
    ).fetchone()['count']

    # توزيع الموظفين على الأقسام
    department_data = conn.execute(
        'SELECT department, COUNT(*) as count FROM employees GROUP BY department'
    ).fetchall()

    conn.close()

    return render_template('dashboard.html',
                           total_employees=total_employees,
                           total_departments=total_departments,
                           total_present=total_present,
                           department_data=department_data)
    
@app.route('/payroll')
def payroll():
    conn = sqlite3.connect('hr.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute('''
        SELECT payrolls.id, employees.name as employee_name, payrolls.salary, 
               payrolls.bonus, payrolls.deductions, payrolls.net_salary, payrolls.payment_date
        FROM payrolls
        JOIN employees ON payrolls.employee_id = employees.id
    ''')
    payrolls = cur.fetchall()
    conn.close()
    return render_template('payroll.html', payrolls=payrolls)

@app.route('/payroll/export/csv')
def payroll_export_csv():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    conn = get_db_connection()

    query = """
        SELECT p.id, e.name AS employee_name, p.salary, p.bonus,
               p.deductions, p.net_salary, p.payment_date
        FROM payrolls p
        JOIN employees e ON p.employee_id = e.id
        ORDER BY p.payment_date DESC, p.id DESC
    """

    rows = conn.execute(query).fetchall()
    conn.close()

    def generate():
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Payroll ID", "Employee", "Salary", "Bonus",
            "Deductions", "Net Salary", "Payment Date"
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        # Data rows (streaming)
        for r in rows:
            writer.writerow([
                r["id"],
                r["employee_name"],
                r["salary"],
                r["bonus"],
                r["deductions"],
                r["net_salary"],
                r["payment_date"]
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=payroll_export.csv"
        }
    )



@app.route('/payroll/add', methods=['GET', 'POST'])
def add_payroll_():
    conn = get_db_connection()

    if request.method == 'POST':
        employee_id = request.form['employee_id']
        salary = float(request.form['salary'])
        payment_date = request.form['payment_date']

        # بنود إضافية (بدلات/خصومات متعددة)
        item_types = request.form.getlist('item_type[]')
        item_labels = request.form.getlist('item_label[]')
        item_amounts = request.form.getlist('item_amount[]')

        allowances_total = 0.0
        deductions_total = 0.0
        items_to_insert = []
        for t, label, amt in zip(item_types, item_labels, item_amounts):
            label = (label or '').strip()
            if not label:
                continue
            try:
                amt_val = float(amt or 0)
            except (TypeError, ValueError):
                amt_val = 0.0
            if amt_val == 0:
                continue
            if t == 'deduction':
                deductions_total += abs(amt_val)
                items_to_insert.append(('deduction', label, abs(amt_val)))
            else:
                allowances_total += abs(amt_val)
                items_to_insert.append(('allowance', label, abs(amt_val)))

        # الحقول القديمة ما زالت مدعومة (bonus/deductions) إن كانت موجودة في الفورم
        bonus = float(request.form.get('bonus', 0) or 0)
        deductions = float(request.form.get('deductions', 0) or 0)
        bonus += allowances_total
        deductions += deductions_total

        net_salary = salary + bonus - deductions

        conn.execute('''
            INSERT INTO payrolls (employee_id, salary, bonus, deductions, net_salary, payment_date)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (employee_id, salary, bonus, deductions, net_salary, payment_date))
        payroll_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        if items_to_insert:
            conn.executemany(
                "INSERT INTO payroll_items (payroll_id, item_type, label, amount) VALUES (?, ?, ?, ?)",
                [(payroll_id, t, label, amt) for (t, label, amt) in items_to_insert]
            )
        conn.commit()
        conn.close()
        return redirect('/payroll')

    keyword = request.args.get('keyword', '').strip()
    if keyword:
        employees = conn.execute(
            """SELECT id, name, email FROM employees
               WHERE name LIKE ? OR COALESCE(email,'') LIKE ? ORDER BY name""",
            ('%' + keyword + '%', '%' + keyword + '%')
        ).fetchall()
    else:
        employees = conn.execute("SELECT id, name, email FROM employees ORDER BY name").fetchall()
    conn.close()
    return render_template('add_payroll.html', employees=employees)

@app.route('/add_payroll', methods=['GET', 'POST'])
def add_payroll():
    conn = get_db_connection()

    if request.method == 'POST':
        employee_id = request.form['employee_id']
        month = request.form.get('month', '')
        salary = request.form.get('salary', '')

        conn.execute("""
            INSERT INTO payrolls (employee_id, month, salary)
            VALUES (?, ?, ?)
        """, (employee_id, month, salary))
        conn.commit()
        conn.close()
        return redirect(url_for('view_payrolls'))

    keyword = request.args.get('keyword', '').strip()
    if keyword:
        employees = conn.execute(
            """SELECT id, name, email FROM employees
               WHERE name LIKE ? OR COALESCE(email,'') LIKE ? ORDER BY name""",
            ('%' + keyword + '%', '%' + keyword + '%')
        ).fetchall()
    else:
        employees = conn.execute("SELECT id, name, email FROM employees ORDER BY name").fetchall()
    conn.close()
    return render_template('add_payroll.html', employees=employees)

@app.route('/payroll')
def view_payrolls():
    conn = sqlite3.connect('hr.db')
    c = conn.cursor()

    # نجيب الرواتب مع أسماء الموظفين
    c.execute("""
        SELECT p.id, e.name, p.month, p.salary
        FROM payrolls p
        JOIN employees e ON p.employee_id = e.id
    """)
    payrolls = c.fetchall()
    conn.close()

    return render_template('payroll.html', payrolls=payrolls)

@app.route('/get_salary/<int:employee_id>')
def get_employee_salary(employee_id):
    conn = sqlite3.connect("hr.db")
    c = conn.cursor()
    c.execute("SELECT salary FROM employees WHERE id = ?", (employee_id,))
    result = c.fetchone()
    conn.close()

    if result:
        return jsonify({"salary": result[0]})
    else:
        return jsonify({"salary": None})

# ========== Routes إدارة الإجازات ==========

@app.route('/leaves')
def leaves():
    """صفحة عرض جميع طلبات الإجازة للمديرين"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    requests = get_all_leave_requests()
    return render_template('leaves.html', requests=requests)

@app.route('/leaves/pending')
def pending_leaves():
    """صفحة عرض طلبات الإجازة المعلقة"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    requests = get_pending_leave_requests()
    return render_template('pending_leaves.html', requests=requests)

@app.route('/leaves/request', methods=['GET', 'POST'])
def request_leave():
    """صفحة طلب إجازة للموظفين"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        employee_id = request.form['employee_id']
        leave_type_id = request.form['leave_type_id']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        reason = request.form['reason']
        
        success, message = submit_leave_request(employee_id, leave_type_id, start_date, end_date, reason)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
        
        return redirect(url_for('request_leave'))
    
    # الحصول على قائمة الموظفين وأنواع الإجازات
    conn = get_db_connection()
    employees = conn.execute('SELECT id, name FROM employees').fetchall()
    leave_types = conn.execute('SELECT id, name, max_days FROM leave_types').fetchall()
    conn.close()
    
    return render_template('request_leave.html', employees=employees, leave_types=leave_types)

@app.route('/leaves/my_requests/<int:employee_id>')
def my_leave_requests(employee_id):
    """صفحة عرض طلبات الإجازة للموظف"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    requests = get_employee_leave_requests(employee_id)
    balances = get_employee_leave_balances(employee_id)
    
    return render_template('my_leave_requests.html', requests=requests, balances=balances, employee_id=employee_id)

@app.route('/leaves/approve/<int:request_id>', methods=['POST'])
def approve_leave(request_id):
    """الموافقة على طلب إجازة"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    approved_by = 1  # يمكن تحديد المدير من الجلسة
    comments = request.form.get('comments', '')
    
    success, message = approve_leave_request(request_id, approved_by, comments)
    
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    
    return redirect(url_for('pending_leaves'))

@app.route('/leaves/reject/<int:request_id>', methods=['POST'])
def reject_leave(request_id):
    """رفض طلب إجازة"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    approved_by = 1  # يمكن تحديد المدير من الجلسة
    comments = request.form.get('comments', '')
    
    success, message = reject_leave_request(request_id, approved_by, comments)
    
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    
    return redirect(url_for('pending_leaves'))

@app.route('/leaves/balance/<int:employee_id>')
def leave_balance(employee_id):
    """عرض رصيد الإجازات للموظف"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # إنشاء رصيد إذا لم يكن موجوداً
    create_leave_balance_for_employee(employee_id)
    
    balances = get_employee_leave_balances(employee_id)
    
    conn = get_db_connection()
    employee = conn.execute('SELECT name FROM employees WHERE id = ?', (employee_id,)).fetchone()
    conn.close()
    
    return render_template('leave_balance.html', balances=balances, employee=employee)

@app.route('/api/leave_balance/<int:employee_id>/<int:leave_type_id>')
def api_leave_balance(employee_id, leave_type_id):
    """API للحصول على رصيد الإجازة"""
    balance = get_employee_leave_balance(employee_id, leave_type_id)
    
    if balance:
        return jsonify({
            'remaining_days': balance['remaining_days'],
            'used_days': balance['used_days'],
            'allocated_days': balance['allocated_days']
        })
    else:
        return jsonify({'error': 'لم يتم العثور على رصيد الإجازة'})

@app.route('/reports/leaves')
def leave_reports():
    """تقارير الإجازات"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    current_year = datetime.now().year
    
    # الحصول على البيانات
    leave_summary = get_annual_leave_summary(current_year)
    leave_usage = get_leave_usage_by_type(current_year)
    
    return render_template('leave_reports.html',
                         leave_summary=leave_summary,
                         leave_usage=leave_usage,
                         year=current_year)


# ========== APIs للتقارير ==========

@app.route('/api/reports/attendance_overview')
def api_attendance_overview():
    """API للحصول على نظرة عامة على الحضور"""
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    overview = get_attendance_overview_by_date(start_date, end_date)
    
    return jsonify([dict(row) for row in overview])

@app.route('/api/reports/payroll_monthly')
def api_payroll_monthly():
    """API للحصول على إجمالي الرواتب الشهرية"""
    year = request.args.get('year', datetime.now().year)
    
    total_payroll = get_total_payroll_by_month(year)
    
    return jsonify([dict(row) for row in total_payroll])

@app.route('/api/reports/leave_usage')
def api_leave_usage():
    """API للحصول على استخدام الإجازات حسب النوع"""
    year = request.args.get('year', datetime.now().year)
    
    usage = get_leave_usage_by_type(year)
    
    return jsonify([dict(row) for row in usage])

@app.route('/api/reports/employee_attendance/<int:employee_id>')
def api_employee_attendance_trend(employee_id):
    """API للحصول على اتجاه حضور موظف معين"""
    year = request.args.get('year', datetime.now().year)
    
    trend = get_employee_attendance_trend(employee_id, year)
    
    return jsonify([dict(row) for row in trend])



# ========== Routes إدارة الأداء ==========

@app.route('/performance')
def performance_dashboard():
    """لوحة معلومات إدارة الأداء الرئيسية"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # الحصول على الفترة النشطة
    active_period = get_active_evaluation_period()
    
    # إحصائيات سريعة
    stats = {}
    if active_period:
        stats['pending_evaluations'] = len(get_pending_evaluations(active_period['id']))
        stats['employees_without_evaluation'] = len(get_employees_without_evaluation(active_period['id']))
        stats['completed_evaluations'] = len(get_evaluations_by_period(active_period['id'])) - stats['pending_evaluations']
    
    return render_template('performance_dashboard.html', 
                         active_period=active_period, 
                         stats=stats)

@app.route('/performance/periods')
def evaluation_periods():
    """إدارة الفترات التقييمية"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    periods = get_evaluation_periods()
    return render_template('evaluation_periods.html', periods=periods)

@app.route('/performance/periods/create', methods=['GET', 'POST'])
def create_period():
    """إنشاء فترة تقييمية جديدة"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        name = request.form['name']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        year = int(request.form['year'])
        quarter = int(request.form['quarter'])
        
        period_id = create_evaluation_period(name, start_date, end_date, year, quarter)
        flash('تم إنشاء الفترة التقييمية بنجاح', 'success')
        return redirect(url_for('evaluation_periods'))
    
    return render_template('create_period.html')

@app.route('/performance/evaluate')
def evaluation_list():
    """قائمة التقييمات"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    period_id = request.args.get('period_id')
    search_query = request.args.get('search', '').strip()
    # إعداد التصفح (Pagination) لقائمة الموظفين
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('page_size', 20))
        if per_page not in (10, 20, 25, 50, 100):
            per_page = 20
    except ValueError:
        per_page = 20
    active_period = get_active_evaluation_period()
    
    if not period_id and active_period:
        period_id = active_period['id']
    
    evaluations = []
    employees = []
    total_employees = 0
    total_employees_without_eval = 0
    pages = 1
    start_index = 0
    end_index = 0
    
    if period_id:
        evaluations = get_evaluations_by_period(period_id)
        conn = get_db_connection()
        params = [period_id]
        base_query = """
            SELECT
                e.id,
                e.name,
                e.department,
                e.position,
                pe.id AS evaluation_id,
                pe.status AS evaluation_status
            FROM employees e
            LEFT JOIN performance_evaluations pe
              ON pe.employee_id = e.id AND pe.period_id = ?
        """
        if search_query:
            base_query += " WHERE e.name LIKE ? OR COALESCE(e.department,'') LIKE ? "
            params.extend([f"%{search_query}%", f"%{search_query}%"])
        base_query += " ORDER BY e.name"

        all_employees = conn.execute(base_query, params).fetchall()
        conn.close()

        total_employees = len(all_employees)
        total_employees_without_eval = sum(1 for row in all_employees if row["evaluation_id"] is None)
        pages = (total_employees + per_page - 1) // per_page if total_employees else 1
        start = (page - 1) * per_page
        end = start + per_page
        employees = all_employees[start:end]
        if total_employees:
            start_index = start + 1
            end_index = min(end, total_employees)
    
    periods = get_evaluation_periods()
    
    return render_template('evaluation_list.html',
                         evaluations=evaluations,
                         employees=employees,
                         total_employees=total_employees,
                         total_employees_in_period=total_employees,
                         total_employees_without_eval=total_employees_without_eval,
                         page=page,
                         pages=pages,
                         periods=periods,
                         search_query=search_query,
                         total_results=total_employees,
                         start_index=start_index,
                         end_index=end_index,
                         selected_period_id=int(period_id) if period_id else None,
                         page_size=per_page)

@app.route('/performance/evaluate/<int:employee_id>')
def evaluate_employee(employee_id):
    """صفحة تقييم موظف"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    period_id = request.args.get('period_id')
    if not period_id:
        active_period = get_active_evaluation_period()
        if active_period:
            period_id = active_period['id']
        else:
            flash('لا توجد فترة تقييمية نشطة', 'error')
            return redirect(url_for('evaluation_list'))
    
    # الحصول على بيانات الموظف
    conn = get_db_connection()
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    period = conn.execute('SELECT * FROM evaluation_periods WHERE id = ?', (period_id,)).fetchone()
    conn.close()
    
    if not employee:
        flash('الموظف غير موجود', 'error')
        return redirect(url_for('evaluation_list'))
    
    # التحقق من وجود تقييم سابق
    existing_evaluation = None
    conn = get_db_connection()
    existing = conn.execute("""
        SELECT * FROM performance_evaluations 
        WHERE employee_id = ? AND period_id = ?
    """, (employee_id, period_id)).fetchone()
    conn.close()
    
    if existing:
        existing_evaluation = get_evaluation_by_id(existing['id'])
    
    criteria = get_evaluation_criteria()
    
    return render_template('evaluate_employee.html',
                         employee=employee,
                         period=period,
                         criteria=criteria,
                         existing_evaluation=existing_evaluation)

@app.route('/performance/evaluate/<int:employee_id>/submit', methods=['POST'])
def submit_evaluation(employee_id):
    """تقديم تقييم الموظف"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    period_id = request.form['period_id']
    evaluator_id = 1  # افتراض أن المستخدم الحالي هو المقيم (يمكن تحسينه)
    
    # جمع بيانات التقييم
    evaluation_data = {
        'overall_rating': request.form['overall_rating'],
        'strengths': request.form['strengths'],
        'areas_for_improvement': request.form['areas_for_improvement'],
        'goals_next_period': request.form['goals_next_period'],
        'comments': request.form['comments'],
        'criteria_details': []
    }
    
    # جمع تفاصيل المعايير
    criteria = get_evaluation_criteria()
    total_score = 0.0
    total_weight = 0.0
    
    for criterion in criteria:
        rating = request.form.get(f'criteria_{criterion["id"]}_rating')
        comments = request.form.get(f'criteria_{criterion["id"]}_comments', '')
        
        if rating:
            evaluation_data['criteria_details'].append({
                'criteria_id': criterion['id'],
                'rating': rating,
                'comments': comments
            })
            
            # حساب النقاط الإجمالية
            score = rating_to_score(rating)
            weight = criterion['weight']
            total_score += score * weight
            total_weight += weight
    
    # حساب النقاط الإجمالية والتقييم العام
    if total_weight > 0:
        overall_score = total_score / total_weight
        evaluation_data['overall_score'] = overall_score
        if not evaluation_data['overall_rating']:
            evaluation_data['overall_rating'] = score_to_rating(overall_score)
    
    try:
        # التحقق من وجود تقييم سابق
        conn = get_db_connection()
        existing = conn.execute("""
            SELECT id FROM performance_evaluations 
            WHERE employee_id = ? AND period_id = ?
        """, (employee_id, period_id)).fetchone()
        conn.close()
        
        if existing:
            # تحديث التقييم الموجود
            update_performance_evaluation(existing['id'], evaluator_id, evaluation_data)
            flash('تم تحديث التقييم بنجاح', 'success')
        else:
            # إنشاء تقييم جديد
            evaluation_id = create_performance_evaluation(employee_id, period_id, evaluator_id, evaluation_data)
            flash('تم إنشاء التقييم بنجاح', 'success')
        
        # إكمال التقييم إذا طُلب ذلك
        if request.form.get('complete_evaluation'):
            if existing:
                complete_performance_evaluation(existing['id'], evaluator_id)
            else:
                complete_performance_evaluation(evaluation_id, evaluator_id)
            flash('تم إكمال التقييم', 'success')
        
    except Exception as e:
        flash(f'حدث خطأ أثناء حفظ التقييم: {str(e)}', 'error')
    
    return redirect(url_for('evaluation_list', period_id=period_id))

@app.route('/performance/view/<int:evaluation_id>')
def view_evaluation(evaluation_id):
    """عرض تفاصيل التقييم"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    evaluation_data = get_evaluation_by_id(evaluation_id)
    
    if not evaluation_data:
        flash('التقييم غير موجود', 'error')
        return redirect(url_for('evaluation_list'))
    
    return render_template('view_evaluation.html', evaluation_data=evaluation_data)

@app.route('/performance/employee/<int:employee_id>/history')
def employee_performance_history(employee_id):
    """تاريخ أداء الموظف"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    # الحصول على بيانات الموظف
    conn = get_db_connection()
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    conn.close()
    
    if not employee:
        flash('الموظف غير موجود', 'error')
        return redirect(url_for('evaluation_list'))
    
    evaluations = get_employee_evaluations(employee_id)
    trend = get_employee_performance_trend(employee_id)
    
    return render_template('employee_performance_history.html',
                         employee=employee,
                         evaluations=evaluations,
                         trend=trend)

@app.route('/performance/reports')
def performance_reports():
    """تقارير الأداء"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    period_id = request.args.get('period_id')
    periods = get_evaluation_periods()
    
    stats = []
    criteria_analysis = []
    
    if period_id:
        stats = get_evaluation_statistics(period_id)
        criteria_analysis = get_criteria_performance_analysis(period_id)
    
    return render_template('performance_reports.html',
                         periods=periods,
                         selected_period_id=int(period_id) if period_id else None,
                         stats=stats,
                         criteria_analysis=criteria_analysis)

@app.route('/performance/employees')
def employees_with_evaluations():
    """صفحة تعرض جميع الموظفين الذين تم تقييمهم عبر كل الفترات"""
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    employees = get_employees_with_evaluations()
    return render_template('employees_with_evaluations.html', employees=employees)

# ========== APIs لإدارة الأداء ==========
@app.template_filter('number_format')
def number_format(value):
    try:
        return "{:,}".format(int(value))
    except (ValueError, TypeError):
        return value
@app.route('/api/performance/criteria')
def api_evaluation_criteria():
    """API للحصول على معايير التقييم"""
    criteria = get_evaluation_criteria()
    return jsonify([dict(row) for row in criteria])

@app.route('/api/performance/periods')
def api_evaluation_periods():
    """API للحصول على الفترات التقييمية"""
    periods = get_evaluation_periods()
    return jsonify([dict(row) for row in periods])

@app.route('/api/performance/statistics/<int:period_id>')
def api_performance_statistics(period_id):
    """API للحصول على إحصائيات الأداء"""
    stats = get_evaluation_statistics(period_id)
    return jsonify([dict(row) for row in stats])

@app.route('/api/performance/employee/<int:employee_id>/trend')
def api_employee_performance_trend(employee_id):
    """API للحصول على اتجاه أداء الموظف"""
    trend = get_employee_performance_trend(employee_id)
    return jsonify([dict(row) for row in trend])

from datetime import datetime

# فلتر تنسيق التاريخ
@app.template_filter('format_date')
def format_date(value):
    if value:
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            return str(value).split(" ")[0]
        except:
            return value
    return "-"

@app.template_filter('number_format')
def number_format_filter(value):
    try:
        return "{:,.0f}".format(float(value))
    except:
        return value
# فلتر تنسيق الوقت
@app.template_filter('format_time')
def format_time(value):
    if value:
        try:
            if isinstance(value, datetime):
                return value.strftime("%H:%M")
            parts = str(value).split(" ")
            if len(parts) > 1:
                return parts[1]
        except:
            pass
    return "-"
from datetime import datetime

@app.template_filter('to_date')
def to_date(value):
    if isinstance(value, str):
        return datetime.strptime(value, "%Y-%m-%d")
    datetime.strptime(value, "%Y-%m-%d %H:%M:%S")

    return value
@app.route('/api/performance/criteria_analysis/<int:period_id>')
def api_criteria_performance_analysis(period_id):
    """API لتحليل الأداء حسب المعايير"""
    analysis = get_criteria_performance_analysis(period_id)
    return jsonify([dict(row) for row in analysis])
@app.route("/export_evaluations/<int:period_id>")
def export_evaluations(period_id):

    conn = sqlite3.connect("hr.db")
    c = conn.cursor()

    c.execute("""
        SELECT employee_name, score, comments
        FROM evaluations
        WHERE period_id = ?
    """, (period_id,))

    data = c.fetchall()
    conn.close()

    return {"data": data}
import os
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
    
